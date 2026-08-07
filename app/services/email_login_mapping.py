from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import (
    AuditLog,
    DismissalSchedule,
    EmailLoginMapping,
    HRSourceRecord,
)
from app.services.ad import ADDirectoryUser, ActiveDirectoryService
from app.services.zimbra import ZimbraAccountIdentity, ZimbraService


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def email_domain(value: str) -> str:
    value = str(value or "").strip().lower()
    return value.rsplit("@", 1)[1] if "@" in value else ""


def normalize_email(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_login(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[\s_\-.]+", "", text)


@dataclass(frozen=True)
class ImportRow:
    row_number: int
    email: str
    login: str


class EmailLoginMappingService:
    """Управляет разделом «Сопоставление e-mail и логина»."""

    LEGACY_SOURCE_IDS = {"org_com"}

    def __init__(self, settings: Settings, db: Session):
        self.settings = settings
        self.db = db

    @property
    def allowed_domains(self) -> tuple[str, ...]:
        return tuple(
            dict.fromkeys(
                domain.strip().lower()
                for domain in self.settings.zimbra_domains
                if domain.strip()
            )
        )

    def resolve_source_domain(self) -> str:
        explicit = self.settings.onec_source_domain.strip().lower()
        if explicit:
            if explicit not in self.allowed_domains:
                raise ValueError(
                    "ONEC_SOURCE_DOMAIN должен быть одним из ZIMBRA_DOMAINS"
                )
            self._migrate_legacy_source(explicit)
            return explicit

        source_ids = {
            str(value or "").strip().lower()
            for value in self.db.scalars(
                select(HRSourceRecord.source_id).where(
                    HRSourceRecord.is_present.is_(True)
                )
            ).all()
            if str(value or "").strip()
        }
        current_domains = source_ids.intersection(self.allowed_domains)
        if len(current_domains) == 1:
            return next(iter(current_domains))

        domains_from_email = {
            email_domain(value)
            for value in self.db.scalars(
                select(HRSourceRecord.corporate_email).where(
                    HRSourceRecord.is_present.is_(True)
                )
            ).all()
            if email_domain(value) in self.allowed_domains
        }
        if len(domains_from_email) == 1:
            domain = next(iter(domains_from_email))
            self._migrate_legacy_source(domain)
            return domain

        if len(self.allowed_domains) == 1:
            domain = self.allowed_domains[0]
            self._migrate_legacy_source(domain)
            return domain

        raise ValueError(
            "Не удалось однозначно определить домен текущей выгрузки 1С. "
            "Задайте ONEC_SOURCE_DOMAIN одним из значений ZIMBRA_DOMAINS."
        )

    def infer_source_domain_from_workbook(self, workbook) -> str:
        explicit = self.settings.onec_source_domain.strip().lower()
        if explicit:
            if explicit not in self.allowed_domains:
                raise ValueError(
                    "ONEC_SOURCE_DOMAIN должен быть одним из ZIMBRA_DOMAINS"
                )
            self._migrate_legacy_source(explicit)
            return explicit

        domains = {
            email_domain(worker.email or "")
            for worker in workbook.workers
            if email_domain(worker.email or "") in self.allowed_domains
        }
        if len(domains) == 1:
            domain = next(iter(domains))
            self._migrate_legacy_source(domain)
            return domain

        if len(self.allowed_domains) == 1:
            domain = self.allowed_domains[0]
            self._migrate_legacy_source(domain)
            return domain

        raise ValueError(
            "В выгрузке 1С не найден один однозначный домен из ZIMBRA_DOMAINS. "
            "Задайте ONEC_SOURCE_DOMAIN."
        )

    def _migrate_legacy_source(self, domain: str) -> None:
        domain = domain.strip().lower()
        if not domain:
            return

        legacy = self.db.scalars(
            select(HRSourceRecord).where(
                HRSourceRecord.source_id.in_(self.LEGACY_SOURCE_IDS)
            )
        ).all()
        if not legacy:
            return

        existing = {
            row.worker_key: row
            for row in self.db.scalars(
                select(HRSourceRecord).where(
                    HRSourceRecord.source_id == domain
                )
            ).all()
        }

        changed = False
        for row in legacy:
            duplicate = existing.get(row.worker_key)
            if duplicate is None:
                row.source_id = domain
                row.source_name = domain
                existing[row.worker_key] = row
            else:
                # A partial new import may already have created the target row.
                # Keep the target and remove the obsolete placeholder row.
                self.db.delete(row)
            changed = True

        if changed:
            self.db.commit()

    def cleanup_dismissed(self, source_domain: str | None = None) -> int:
        domain = source_domain or self.resolve_source_domain()
        today = date.today()

        schedules = self.db.scalars(
            select(DismissalSchedule).where(
                DismissalSchedule.dismissal_date <= today,
                DismissalSchedule.ad_expiration_set.is_(True),
                DismissalSchedule.zimbra_note_set.is_(True),
            )
        ).all()

        due = [
            item
            for item in schedules
            if email_domain(item.corporate_email) == domain
        ]
        if not due:
            return 0

        mappings = self.db.scalars(
            select(EmailLoginMapping).where(
                EmailLoginMapping.source_domain == domain
            )
        ).all()

        removed = 0
        for mapping in mappings:
            for schedule in due:
                same_login = (
                    mapping.ad_login.strip().lower()
                    == schedule.login.strip().lower()
                )
                same_email = schedule.corporate_email.strip().lower() in {
                    mapping.source_email.strip().lower(),
                    mapping.zimbra_email.strip().lower(),
                }
                if same_login or same_email:
                    self.db.delete(mapping)
                    removed += 1
                    break
        if removed:
            self.db.commit()
        return removed

    def list_mappings(self, source_domain: str | None = None) -> list[dict]:
        domain = source_domain or self.resolve_source_domain()
        self.cleanup_dismissed(domain)

        records = {
            row.worker_key: row
            for row in self.db.scalars(
                select(HRSourceRecord).where(
                    HRSourceRecord.source_id == domain
                )
            ).all()
        }

        mappings = self.db.scalars(
            select(EmailLoginMapping)
            .where(EmailLoginMapping.source_domain == domain)
            .order_by(EmailLoginMapping.source_email)
        ).all()

        rows = []
        for mapping in mappings:
            record = records.get(mapping.worker_key)
            rows.append(
                {
                    "id": mapping.id,
                    "worker_key": mapping.worker_key,
                    "fio": record.fio if record else "Работник отсутствует в текущей выгрузке",
                    "source_email": mapping.source_email,
                    "ad_login": mapping.ad_login,
                    "zimbra_email": mapping.zimbra_email,
                    "ad_object_guid": mapping.ad_object_guid,
                    "zimbra_id": mapping.zimbra_id,
                    "last_verified_at": mapping.last_verified_at,
                    "present": bool(record and record.is_present),
                }
            )
        return rows

    def mapping_by_worker(
        self,
        worker_keys: list[str],
        source_domain: str,
    ) -> dict[str, EmailLoginMapping]:
        if not worker_keys:
            return {}
        return {
            item.worker_key: item
            for item in self.db.scalars(
                select(EmailLoginMapping).where(
                    EmailLoginMapping.source_domain == source_domain,
                    EmailLoginMapping.worker_key.in_(worker_keys),
                )
            ).all()
        }

    def _record_by_email(
        self,
        source_email: str,
        source_domain: str,
    ) -> HRSourceRecord:
        normalized = normalize_email(source_email)
        records = self.db.scalars(
            select(HRSourceRecord).where(
                HRSourceRecord.source_id == source_domain,
                HRSourceRecord.is_present.is_(True),
                HRSourceRecord.corporate_email == normalized,
            )
        ).all()
        if not records:
            raise ValueError(
                f"{normalized}: e-mail не найден среди работников текущей выгрузки 1С"
            )
        if len(records) > 1:
            raise ValueError(
                f"{normalized}: найдено несколько работников, требуется ручная проверка"
            )
        return records[0]

    def _save_mapping(
        self,
        *,
        record: HRSourceRecord,
        source_domain: str,
        source_email: str,
        ad_user: ADDirectoryUser,
        zimbra: ZimbraAccountIdentity,
        actor: str,
    ) -> tuple[str, EmailLoginMapping | None]:
        if not ad_user.object_guid:
            raise ValueError(
                f"{ad_user.username}: AD не вернул objectGUID"
            )
        if not zimbra.zimbra_id:
            raise ValueError(
                f"{source_email}: Zimbra не вернула zimbraId"
            )

        # If the actual AD login and actual Zimbra primary login are equal,
        # the record is no longer an exception.
        if ad_user.username.lower() == zimbra.login.lower():
            old = self.db.scalar(
                select(EmailLoginMapping).where(
                    EmailLoginMapping.worker_key == record.worker_key,
                    EmailLoginMapping.source_domain == source_domain,
                )
            )
            if old is not None:
                self.db.delete(old)
                self.db.add(
                    AuditLog(
                        actor=actor,
                        action="email_login_mapping_auto_remove",
                        target=source_email,
                        result="success",
                        details="Логины AD и Zimbra стали одинаковыми",
                    )
                )
                self.db.commit()
            return "not_needed", None

        mapping = self.db.scalar(
            select(EmailLoginMapping).where(
                EmailLoginMapping.worker_key == record.worker_key,
                EmailLoginMapping.source_domain == source_domain,
            )
        )
        created = mapping is None
        if mapping is None:
            mapping = EmailLoginMapping(
                worker_key=record.worker_key,
                source_domain=source_domain,
                source_email=source_email,
                ad_object_guid=ad_user.object_guid,
                ad_login=ad_user.username,
                zimbra_id=zimbra.zimbra_id,
                zimbra_email=zimbra.primary_email,
                created_by=actor,
                last_verified_at=utcnow(),
            )
            self.db.add(mapping)
        else:
            mapping.source_email = source_email
            mapping.ad_object_guid = ad_user.object_guid
            mapping.ad_login = ad_user.username
            mapping.zimbra_id = zimbra.zimbra_id
            mapping.zimbra_email = zimbra.primary_email
            mapping.last_verified_at = utcnow()

        self.db.add(
            AuditLog(
                actor=actor,
                action=(
                    "email_login_mapping_create"
                    if created
                    else "email_login_mapping_update"
                ),
                target=source_email,
                result="success",
                details=f"AD={ad_user.username}; Zimbra={zimbra.primary_email}",
            )
        )
        self.db.commit()
        self.db.refresh(mapping)
        return ("created" if created else "updated"), mapping

    def save_confirmed_identity(
        self,
        *,
        record: HRSourceRecord,
        ad_user: ADDirectoryUser,
        zimbra: ZimbraAccountIdentity,
        actor: str,
    ) -> dict:
        """Сохранить подтвержденную оператором связь без повторного поиска."""
        source_email = normalize_email(record.corporate_email)
        if not source_email or "@" not in source_email:
            raise ValueError("У работника нет корректного корпоративного e-mail")

        source_domain = str(record.source_id or "").strip().lower()
        if source_domain not in self.allowed_domains:
            source_domain = self.resolve_source_domain()

        if email_domain(source_email) != source_domain:
            raise ValueError(
                f"Для текущей выгрузки ожидается домен {source_domain}"
            )

        status, saved = self._save_mapping(
            record=record,
            source_domain=source_domain,
            source_email=source_email,
            ad_user=ad_user,
            zimbra=zimbra,
            actor=actor,
        )
        return {
            "status": status,
            "mapping_id": saved.id if saved else None,
            "fio": record.fio,
            "email": source_email,
            "ad_login": ad_user.username,
            "zimbra_login": zimbra.login,
        }


    def add_manual(
        self,
        source_email: str,
        ad_login: str,
        actor: str,
    ) -> dict:
        domain = self.resolve_source_domain()
        source_email = normalize_email(source_email)
        ad_login = normalize_login(ad_login)
        if not source_email or "@" not in source_email:
            raise ValueError("Укажите корректный e-mail")
        if email_domain(source_email) != domain:
            raise ValueError(
                f"Для текущей выгрузки ожидается домен {domain}"
            )
        if not ad_login:
            raise ValueError("Укажите логин AD")

        record = self._record_by_email(source_email, domain)
        ad_user = ActiveDirectoryService(self.settings).get_user(ad_login)
        if ad_user is None:
            raise ValueError(f"AD: логин {ad_login} не найден")

        zimbra = ZimbraService(self.settings).account_by_address(source_email)
        if zimbra is None:
            raise ValueError(f"Zimbra: адрес {source_email} не найден")

        status, mapping = self._save_mapping(
            record=record,
            source_domain=domain,
            source_email=source_email,
            ad_user=ad_user,
            zimbra=zimbra,
            actor=actor,
        )
        return {
            "status": status,
            "mapping_id": mapping.id if mapping else None,
            "fio": record.fio,
            "email": source_email,
            "ad_login": ad_user.username,
            "zimbra_login": zimbra.login,
        }

    @staticmethod
    def parse_xlsx(data: bytes) -> list[ImportRow]:
        if not data:
            raise ValueError("Файл пуст")
        if len(data) > 10 * 1024 * 1024:
            raise ValueError("Файл больше 10 МБ")

        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
        try:
            sheet = workbook.active
            email_col = None
            login_col = None
            header_row = None

            email_headers = {
                "email",
                "emailадрес",
                "адресэлектроннойпочты",
                "электроннаяпочта",
                "почта",
            }
            login_headers = {
                "логин",
                "login",
                "логинад",
                "adlogin",
                "samaccountname",
            }

            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                normalized = [
                    normalize_header(sheet.cell(row_idx, col).value)
                    for col in range(1, sheet.max_column + 1)
                ]
                for col_idx, header in enumerate(normalized, start=1):
                    if header in email_headers and email_col is None:
                        email_col = col_idx
                    if header in login_headers and login_col is None:
                        login_col = col_idx
                if email_col and login_col:
                    header_row = row_idx
                    break

            if not header_row:
                raise ValueError(
                    "В XLSX не найдены колонки «e-mail» и «логин»"
                )

            rows: list[ImportRow] = []
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                email = normalize_email(
                    sheet.cell(row_idx, email_col).value
                )
                login = normalize_login(
                    sheet.cell(row_idx, login_col).value
                )
                if not email and not login:
                    continue
                rows.append(
                    ImportRow(
                        row_number=row_idx,
                        email=email,
                        login=login,
                    )
                )
            return rows
        finally:
            workbook.close()

    def import_xlsx(
        self,
        data: bytes,
        actor: str,
    ) -> dict:
        domain = self.resolve_source_domain()
        rows = self.parse_xlsx(data)
        if not rows:
            return {
                "total": 0,
                "created": 0,
                "updated": 0,
                "not_needed": 0,
                "errors": [],
            }

        records = self.db.scalars(
            select(HRSourceRecord).where(
                HRSourceRecord.source_id == domain,
                HRSourceRecord.is_present.is_(True),
            )
        ).all()
        by_email: dict[str, list[HRSourceRecord]] = {}
        for record in records:
            by_email.setdefault(
                normalize_email(record.corporate_email),
                [],
            ).append(record)

        logins = sorted(
            {row.login for row in rows if row.login}
        )
        emails = sorted(
            {
                row.email
                for row in rows
                if row.email and email_domain(row.email) == domain
            }
        )

        ad_error = ""
        zimbra_error = ""
        ad_users: dict[str, ADDirectoryUser] = {}
        zimbra_accounts: dict[str, ZimbraAccountIdentity] = {}

        try:
            ad_users = ActiveDirectoryService(
                self.settings
            ).users_by_logins(logins)
        except Exception as exc:
            ad_error = str(exc)

        try:
            zimbra_accounts = ZimbraService(
                self.settings
            ).accounts_by_addresses(emails)
        except Exception as exc:
            zimbra_error = str(exc)

        result = {
            "total": len(rows),
            "created": 0,
            "updated": 0,
            "not_needed": 0,
            "errors": [],
        }

        for row in rows:
            try:
                if not row.email or "@" not in row.email:
                    raise ValueError("не указан корректный e-mail")
                if email_domain(row.email) != domain:
                    raise ValueError(
                        f"e-mail не относится к домену {domain}"
                    )
                if not row.login:
                    raise ValueError("не указан логин")
                if ad_error:
                    raise ValueError(f"AD: {ad_error}")
                if zimbra_error:
                    raise ValueError(f"Zimbra: {zimbra_error}")

                record_matches = by_email.get(row.email, [])
                if not record_matches:
                    raise ValueError(
                        "e-mail не найден в текущей выгрузке 1С"
                    )
                if len(record_matches) > 1:
                    raise ValueError(
                        "e-mail относится к нескольким работникам"
                    )

                ad_user = ad_users.get(row.login)
                if ad_user is None:
                    raise ValueError("логин не найден в AD")
                zimbra = zimbra_accounts.get(row.email)
                if zimbra is None:
                    raise ValueError("e-mail не найден в Zimbra")

                status, _ = self._save_mapping(
                    record=record_matches[0],
                    source_domain=domain,
                    source_email=row.email,
                    ad_user=ad_user,
                    zimbra=zimbra,
                    actor=actor,
                )
                result[status] += 1
            except Exception as exc:
                result["errors"].append(
                    {
                        "row": row.row_number,
                        "email": row.email,
                        "login": row.login,
                        "error": str(exc),
                    }
                )

        return result

    def delete_mapping(
        self,
        mapping_id: int,
        actor: str,
    ) -> None:
        mapping = self.db.get(EmailLoginMapping, mapping_id)
        if mapping is None:
            raise ValueError("Сопоставление не найдено")

        target = mapping.source_email
        self.db.delete(mapping)
        self.db.add(
            AuditLog(
                actor=actor,
                action="email_login_mapping_delete",
                target=target,
                result="success",
            )
        )
        self.db.commit()
